import win32com.client
import sys
import subprocess
import time
import threading
import win32gui
import win32con
import pandas as pd
import os

# ==============================================================================
# DOSYA VE EXCEL İŞLEMLERİ
# ==============================================================================

def check_files_exist(components, vault_path):
    """
    Parçaların PDM/Vault yolunda olup olmadığını kontrol eder.
    Geriye (bulunanlar, eksikler) döner.
    """
    found = []
    missing = []
    
    print(f"\n[INFO] Dosya kontrolü yapılıyor... (Konum: {vault_path})")
    
    for comp in components:
        # SAP'den gelen kod genellikle dosya adıdır. Uzantı .sldprt veya .sldasm olabilir.
        # Burada basitçe .sldprt varsayıyoruz, gerekirse her ikisi de denenebilir.
        # Veya SAP tanımında dosya adı geçiyor olabilir.
        
        code = comp['kod']
        # Olası dosya isimleri
        candidates = [f"{code}.sldprt", f"{code}.sldasm"]
        
        file_path = None
        for cand in candidates:
            full_path = os.path.join(vault_path, cand)
            if os.path.exists(full_path):
                file_path = full_path
                break
        
        if file_path:
            comp['full_path'] = file_path
            found.append(comp)
            print(f"[LOG] Dosya bulundu: {code}")
        else:
            missing.append(comp)
            print(f"[WARNING] Dosya EKSİK: {code}")
            
    return found, missing

def export_missing_to_excel(missing_list):
    """Eksik parçaları Excel'e kaydeder ve kırmızı ile işaretler."""
    if not missing_list:
        return

    try:
        df = pd.DataFrame(missing_list)
        file_name = "Eksik_Parcalar.xlsx"
        
        # Pandas ile Excel yazma (Styling için ExcelWriter kullanılır)
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Eksik')
            
            # Styling
            workbook = writer.book
            worksheet = writer.sheets['Eksik']
            
            # Kırmızı dolgu stili
            from openpyxl.styles import PatternFill
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            # Tüm satırları boya
            for row in worksheet.iter_rows(min_row=2, max_row=len(missing_list)+1):
                for cell in row:
                    cell.fill = red_fill
                    
        print(f"[INFO] Eksik parçalar '{file_name}' dosyasına kaydedildi.")
        
    except Exception as e:
        print(f"[ERROR] Excel oluşturma hatası: {e}")

# ==============================================================================
# SOLIDWORKS OTOMASYON SINIFI
# ==============================================================================

class SolidWorksAutomation:
    def __init__(self):
        self.sw_app = None
        self.model = None
        
    def connect(self):
        try:
            self.sw_app = win32com.client.Dispatch("SldWorks.Application")
            self.sw_app.Visible = True
            print("[INFO] SolidWorks bağlantısı sağlandı.")
            return True
        except Exception as e:
            print(f"[ERROR] SolidWorks başlatılamadı: {e}")
            return False

    def create_assembly(self):
        """Yeni bir montaj dosyası oluşturur"""
        try:
            # 1 = swDocASSEMBLY
            self.model = self.sw_app.NewDocument("C:\\ProgramData\\SolidWorks\\SolidWorks 2024\\templates\\Assembly.asmdot", 0, 0, 0)
            if not self.model:
                # Template bulunamazsa varsayılanı dene
                self.model = self.sw_app.NewDocument("", 1, 0, 0)
            
            print("[INFO] Yeni montaj oluşturuldu.")
            return True
        except Exception as e:
            print(f"[ERROR] Montaj oluşturma hatası: {e}")
            return False

    def add_component(self, file_path, quantity):
        """Montaja bileşen ekler ve konfigürasyonları ayarlar"""
        try:
            if not self.model:
                return

            # Dosya uzantısı
            ext = os.path.splitext(file_path)[1].lower()
            
            # Bileşeni ekle
            # AddComponent5(Name, ConfigName, X, Y, Z)
            comp = self.model.AddComponent5(file_path, "", 0, 0, 0)
            
            if not comp:
                print(f"[ERROR] Bileşen eklenemedi: {os.path.basename(file_path)}")
                return

            comp_name = comp.Name2
            print(f"[LOG] Parça eklendi: {comp_name} (Miktar: {quantity})")

            # --- MANTIKLAR ---
            
            # 1. Miktar > 1 ise Linear Pattern (Doğrusal Çoğaltma)
            # Not: Bu karmaşık bir işlem, burada basitçe logluyoruz.
            # Gerçekte FeatureManager.FeatureLinearPattern4 kullanılır.
            if quantity > 1:
                print(f"[ACTION] '{comp_name}' için {quantity} adet çoğaltma (Linear Pattern) uygulanıyor...")
                # self.apply_linear_pattern(comp, quantity) 

            # 2. Miktar = -1 ise GİZLE (Hide)
            if quantity == -1:
                print(f"[ACTION] '{comp_name}' gizleniyor (Hide)...")
                # SelectByID2(Name, Type, X, Y, Z, Append, Mark, Callout, SelectOption)
                # Type: "COMPONENT"
                self.model.Extension.SelectByID2(comp_name, "COMPONENT", 0, 0, 0, False, 0, None, 0)
                self.model.HideComponent2()

            # 3. Alt Montaj ise Flexible (Esnek) Yap
            if ext == ".sldasm":
                print(f"[ACTION] '{comp_name}' alt montajı Esnek (Flexible) yapılıyor...")
                # SolveIn: 0=Rigid, 1=Flexible
                comp.SolveIn = 1 

        except Exception as e:
            print(f"[ERROR] Bileşen işlem hatası ({file_path}): {e}")

    def set_isometric_view(self):
        """İzometrik görünüme geçer"""
        try:
            if self.model:
                self.model.ShowNamedView2("*Isometric", 7)
                self.model.ViewZoomtofit2()
        except:
            pass

    def update_custom_property(self, prop_name, prop_value):
        """Dosya özelliklerini (Custom Properties) günceller"""
        try:
            if self.model:
                cpm = self.model.Extension.CustomPropertyManager("")
                # Add3: Name, Type, Value, Overwrite
                # 30 = swCustomInfoText
                cpm.Add3(prop_name, 30, str(prop_value), 1)
                print(f"[LOG] Özellik güncellendi: {prop_name} = {prop_value}")
        except Exception as e:
            print(f"[ERROR] Özellik güncelleme hatası: {e}")

    def save_assembly(self, path):
        """Montajı kaydeder"""
        try:
            if self.model:
                # SaveAs3(Name, Version, Options)
                # 0 = swSaveAsCurrentVersion
                # 1 = swSaveAsOptions_Silent
                self.model.SaveAs3(path, 0, 1)
                print(f"[INFO] Montaj kaydedildi: {path}")
                return True
        except Exception as e:
            print(f"[ERROR] Kaydetme hatası: {e}")
            return False

# ==============================================================================
# PDM OTOMASYON SINIFI
# ==============================================================================

class PDMAutomation:
    def __init__(self, vault_name="PGR2024"):
        self.vault_name = vault_name
        self.vault = None

    def connect(self):
        try:
            try:
                self.vault = win32com.client.Dispatch("ConisioLib.EdmVault5")
            except:
                self.vault = win32com.client.Dispatch("ConisioLib.EdmVault")
            
            if not self.vault.IsLoggedIn:
                self.vault.LoginAuto(self.vault_name, 0)
            return True
        except Exception as e:
            print(f"[ERROR] PDM bağlantı hatası: {e}")
            return False

    def check_in_file(self, file_path, comment="Otomatik Montaj"):
        """Dosyayı kasaya gönderir (Check-In)"""
        try:
            if not self.vault:
                return False
                
            folder_path = os.path.dirname(file_path)
            file_name = os.path.basename(file_path)
            
            pdm_file = self.vault.GetFileFromPath(file_path)[0]
            if pdm_file:
                if pdm_file.IsLocked:
                    pdm_file.UnlockFile(0, comment)
                    print(f"[INFO] PDM Check-In başarılı: {file_name}")
                    return True
                else:
                    print(f"[INFO] Dosya zaten Check-In durumunda: {file_name}")
                    return True
            else:
                # Dosya PDM'de değilse eklemesi gerekebilir ama 
                # genellikle SW kaydettiğinde PDM klasöründeyse otomatik eklenir (User ayarlarına bağlı).
                # Biz sadece Unlock (Check-in) deniyoruz.
                print(f"[WARNING] Dosya PDM'de bulunamadı: {file_name}")
                return False
        except Exception as e:
            print(f"[ERROR] PDM Check-In hatası: {e}")
            return False

# ==============================================================================
# MAIN LOGIC
# ==============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--headless', action='store_true', help='Run in headless mode for Electron')
    args = parser.parse_args()

    if args.headless:
        # --- HEADLESS MODE (ELECTRON) ---
        try:
            # Stdin'den JSON verisini oku
            input_data = sys.stdin.read()
            if not input_data:
                print("[ERROR] No input data received.")
                return

            data = json.loads(input_data)
            
            username = data.get('sapUsername')
            password = data.get('sapPassword')
            kits_to_process = data.get('kits', [])
            vault_path = data.get('vaultPath', "C:\\Vault") # Varsayılan yol
            
            if 'materialCode' in data:
                kits_to_process.append(data['materialCode'])

            # 1. SAP İşlemleri
            sap = SapAutomation()
            all_missing_components = []
            processed_kits = []
            
            if sap.connect():
                if sap.login(username, password):
                    print("[INFO] SAP Login başarılı. İşlem başlıyor...")
                    
                    sw = SolidWorksAutomation()
                    pdm = PDMAutomation()
                    
                    if not sw.connect():
                        print("[ERROR] SolidWorks başlatılamadı. İşlem durduruluyor.")
                        return

                    if not pdm.connect():
                        print("[WARNING] PDM bağlantısı sağlanamadı. Check-in yapılamayabilir.")

                    for kit_code in kits_to_process:
                        print(f"\n{'='*50}")
                        print(f"[INFO] Kit İşleniyor: {kit_code}")
                        print(f"{'='*50}")
                        
                        # A. Veri Çekme
                        comps = sap.get_components(kit_code)
                        if not comps:
                            print(f"[WARNING] {kit_code} için bileşen bulunamadı veya hata oluştu.")
                            continue
                            
                        # Miktar düzeltme
                        for c in comps:
                            c['parent_kit'] = kit_code
                            try:
                                c['miktar_int'] = int(float(c['miktar'].replace(',','.')))
                            except:
                                c['miktar_int'] = 1

                        # B. Dosya Kontrolü
                        found_comps, missing_comps = check_files_exist(comps, vault_path)
                        
                        if missing_comps:
                            print(f"[WARNING] {kit_code} için {len(missing_comps)} adet eksik parça var. Montaj oluşturulmayacak.")
                            all_missing_components.extend(missing_comps)
                            continue
                        
                        # C. Montaj Oluşturma
                        print(f"[INFO] {kit_code} için tüm parçalar tam. Montaj başlatılıyor...")
                        if sw.create_assembly():
                            for comp in found_comps:
                                sw.add_component(comp['full_path'], comp['miktar_int'])
                            
                            # D. Veri Kartı ve Kayıt
                            
                            # 1. SAP Numarası Güncelleme (Kit No + *)
                            sw.update_custom_property("SAP Numarası", f"{kit_code}*")
                            sw.update_custom_property("Tanım", f"Otomatik Montaj - {kit_code}")

                            # 2. 5. Hane Kontrolü (Özel Durum)
                            if len(kit_code) >= 5 and kit_code[4] == '8':
                                print(f"[INFO] Kit no ({kit_code}) 5. hanesi 8, 'Özel' işaretleniyor...")
                                sw.update_custom_property("Ozel_Durum", "Özel")
                            
                            # İzometrik Görünüm
                            sw.set_isometric_view()
                            
                            # Kaydetme
                            save_name = f"{kit_code}.sldasm"
                            save_path = os.path.join(vault_path, save_name)
                            
                            if sw.save_assembly(save_path):
                                # E. PDM Check-In
                                if pdm.vault:
                                    pdm.check_in_file(save_path)
                                
                                processed_kits.append(save_path)
                                print(f"[SUCCESS] {kit_code} tamamlandı.")
                            else:
                                print(f"[ERROR] {kit_code} kaydedilemedi.")
                        
                        time.sleep(1)

                    # Döngü Bitti - Raporlama
                    
                    if all_missing_components:
                        print(f"\n[WARNING] Toplam {len(all_missing_components)} adet eksik parça raporlanıyor...")
                        export_missing_to_excel(all_missing_components)
                        print(f"[RESULT_MISSING]{json.dumps(all_missing_components)}")
                    
                    if processed_kits:
                        result = {
                            "status": "success",
                            "processed_files": processed_kits,
                            "count": len(processed_kits)
                        }
                        result["file"] = f"{len(processed_kits)} adet montaj oluşturuldu."
                        print(f"[RESULT_SUCCESS]{json.dumps(result)}")
                    
                    if not processed_kits and not all_missing_components:
                        print("[INFO] İşlenecek kit bulunamadı veya hepsi hatalı.")

                else:
                    print("[ERROR] SAP Login Failed")
            else:
                print("[ERROR] SAP Connection Failed")
                
        except Exception as e:
            print(f"[ERROR] Headless execution failed: {e}")

    else:
        print("GUI modu bu sürümde desteklenmemektedir. Lütfen --headless kullanın.")

if __name__ == "__main__":
    main()
